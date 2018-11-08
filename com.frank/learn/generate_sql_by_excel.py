# coding=utf-8
import pymysql
from openpyxl import load_workbook
import json



def get_city_code(s):
    print("s=", s)
    if s.find('东城') != -1:
        return '110101'
    elif s.find('西城') != -1:
        return '110102'
    elif s.find('通州') != -1:
        return '110112'
    elif s.find('朝阳') != -1:
        return '110105'
    elif s.find('昌平') != -1:
        return '110114'
    elif s.find('大兴') != -1:
        return '110115'
    elif s.find('海淀') != -1:
        return '110108'
    elif s.find('密云') != -1:
        return '110228'
    elif s.find('丰台') != -1:
        return '110106'
    else:
        return '123'


def get_bank_code(s):
    if s.find('工商银行') != -1:
        return 'ICBC'
    elif s.find('中国银行') != -1:
        return 'BOC'
    elif s.find('建设银行') != -1:
        return 'CCB'
    elif s.find('邮政储蓄') != -1:
        return 'POST'
    elif s.find('中信银行') != -1:
        return 'ECITIC'
    elif s.find('光大银行') != -1:
        return 'CEB'
    elif s.find('民生银行') != -1:
        return 'CMBC'
    elif s.find('广发银行') != -1:
        return 'GDB'
    elif s.find('招商银行') != -1:
        return 'CMBCHINA'
    elif s.find('兴业银行') != -1:
        return 'CIB'
    elif s.find('平安银行') != -1:
        return 'PINGAN'
    elif s.find('华夏银行') != -1:
        return 'HXB'
    elif s.find('上海浦东发展银行') != -1:
        return 'SPDB'
    elif s.find('农业银行') != -1:
        return 'ABC'
    elif s.find('交通银行') != -1:
        return 'BOCO'
    elif s.find('北京银行') != -1:
        return 'BCCB'
    else:
        return '123'


def generate_sql(file_path):
    connect = pymysql.connect(host='192.168.xxx', port=1234, user='111', passwd='111', db='test')
    # connect.set_character_set('utf8')
    cursor = connect.cursor()
    cursor.execute("select version()")
    fetchone = cursor.fetchone()
    print("fetchone=", fetchone)

    print("row=", file_path)
    wb = load_workbook(file_path)
    sheet = wb.get_sheet_by_name('Sheet1')
    print("sheet=", sheet)
    # rows = sheet.iter_rows(range_string=None, row_offset=0, column_offset=0)

    # for i in range(1, 4):
    #     for j in range(1, 3):
    #         print(sheet.cell(row=i, column=j).value)
    i = 0
    perfix = 'TEST_00000'
    id = 100

    for row in sheet.rows:
        shop_name = row[0].value
        shop_code = perfix + str(i)
        i = i + 1
        id = id + 1
        city = get_city_code(row[1].value)

        data = {
            'area': row[2].value,
            'duration': row[4].value,
            'number': row[7].value
        }

        # ensure_ascii=False 要不汉字的编码入库有问题
        extra_info = json.dumps(data, ensure_ascii=False)

        insert_sql = "INSERT INTO `your_table` ( `id`,``, ``, ``, ``, ``, ``, ``, ``, ``, ``) " \
                     "VALUES ( '{}','1', '{}', '{}', '110000', '{}', '{}', '', '{}', '1', '{}');"
        _sql = insert_sql.format(id, shop_code, shop_name, city, shop_name, extra_info)

        print("_sql=", _sql)

        # .encode("utf-8")
        execute = cursor.execute(_sql.encode("utf-8"))
        print("execute=", execute)
        # 最后插入行的主键id
        print("ID of last record is ", int(cursor.lastrowid))  # 最后插入行的主键ID
        connect.commit()

print("zhixing", generate_sql('C:/Users/test2.xlsx'))

