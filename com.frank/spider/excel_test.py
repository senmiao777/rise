# 读写2007 excel
import openpyxl
from datetime import datetime


def write07Excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '2007测试表'

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")


def write07Excel2(txt_path, excel_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '2018.05.31数据'

    with open(file=txt_path, mode="r", encoding="utf-8") as f:
        # 为a+模式时，因为为追加模式，指针已经移到文尾，读出来的是一个空字符串。
        #   ftext = f.read()  # 一次性读全部成一个字符串
        ftextlist = f.readlines()  # 也是一次性读全部，但每一行作为一个子句存入一个列表
        lines = len(ftextlist)
        print("lines =", lines)
        for i in range(lines):
            split = ftextlist[i].split(",")
            created_at = split[0].strip()
            updated_at = split[1].strip()
            sheet.cell(row=i + 1, column=1, value=created_at)
            sheet.cell(row=i + 1, column=2, value=updated_at)
            status_desc = get_status_desc(int(split[2].strip()))
            sheet.cell(row=i + 1, column=3, value=status_desc)
            strptime = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
            strptime2 = datetime.strptime(updated_at, "%Y-%m-%d %H:%M:%S")
            sheet.cell(row=i + 1, column=4, value=(strptime2 - strptime))

    wb.save(excel_path)
    print("写入数据成功！")


def get_status_desc(status):
    if status == 3:
        return "分配"
    elif status == 1:
        return "分配中"
    else:
        return "未知状态"


print("写入Excel ", write07Excel2("E:/test2.txt","E:/data3.xlsx"))
